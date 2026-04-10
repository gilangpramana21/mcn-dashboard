"""
Report Export Service — Generate Excel 1 (Outreach), Excel 2 (Deal), Excel 3 (Master Brand)
"""
from __future__ import annotations
import io
from typing import Optional
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ─── Header definitions ───────────────────────────────────────────────────────

EXCEL1_HEADERS = [
    "TANGGAL", "USERNAME", "LINK ACC", "FOLLS", "CONTACT (WA)", "BRAND", "PIC",
    "AVG GMV/MONTH", "GMV PER PEMBELI", "UPDATE", "RESPON", "SPEED", "RESULT",
    "SAMPEL GRATIS", "NOTE", "ID PESANAN", "NO VA C.O SAMPEL",
    "STATUS PAYMENT SAMPEL", "GMV PER WEEK 1 AFTER JOIN", "GMV PER WEEK 2 AFTER JOIN",
    "GMV PER WEEK 3 AFTER JOIN", "GMV PER WEEK 4 AFTER JOIN", "GMV PERBULAN AFTER JOIN"
]

EXCEL2_HEADERS = [
    "TANGGAL", "USERNAME", "LINK ACC", "FOLLS", "CONTACT (WA)", "BRAND", "PIC",
    "AVG GMV/MONTH", "GMV PER PEMBELI", "UPDATE", "RESPON", "SPEED", "RESULT",
    "STATUS SEMPEL", "LINK VIDEO", "TOTAL VT", "NOTE DEAL", "NOTE DARI RARA",
    "ID PESANAN", "NO VA C.O SAMPEL", "STATUS PAYMENT SAMPEL",
    "GMV PER WEEK 1 AFTER JOIN", "GMV PER WEEK 2 AFTER JOIN",
    "GMV PER WEEK 3 AFTER JOIN", "GMV PER WEEK 4 AFTER JOIN", "GMV PERBULAN AFTER JOIN"
]

EXCEL3_HEADERS = [
    "BRAND", "NAMA PRODUK", "LINK SKU AFFILIASI", "HARGA", "NO WA YANG DIPAKAI", "SOW"
]


def _apply_header_style(ws, headers: list, fill_color: str = "FFD700"):
    """Apply header styling."""
    if not OPENPYXL_AVAILABLE:
        return
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="000000")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 15)


async def generate_excel1_outreach(
    db: AsyncSession,
    brand_id: Optional[str] = None,
) -> bytes:
    """Generate Excel 1 — Database Outreach (sheet per brand)."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl tidak terinstall. Jalankan: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Ambil semua brand yang punya data outreach
    brand_query = """
        SELECT DISTINCT b.id, b.name
        FROM influencers i
        JOIN brands b ON i.brand_id = b.id
        ORDER BY b.name
    """
    params = {}
    if brand_id:
        brand_query = "SELECT id, name FROM brands WHERE id = :brand_id"
        params["brand_id"] = brand_id

    brand_result = await db.execute(text(brand_query), params)
    brands = brand_result.fetchall()

    if not brands:
        ws = wb.create_sheet("DATABASE OUTREACH")
        _apply_header_style(ws, EXCEL1_HEADERS, "FFD700")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    for brand_row in brands:
        sheet_name = str(brand_row[1])[:31]
        ws = wb.create_sheet(sheet_name)
        _apply_header_style(ws, EXCEL1_HEADERS, "FFD700")

        query = """
            SELECT
                i.created_at::date as tanggal,
                i.tiktok_user_id as username,
                CONCAT('tiktok.com/@', REPLACE(COALESCE(i.tiktok_user_id, ''), '@', ''), '/shop') as link_acc,
                i.follower_count as folls,
                i.phone_number as contact_wa,
                b.name as brand,
                i.pic,
                i.avg_gmv_per_month,
                i.gmv_per_buyer,
                i.update_status,
                i.respon_status,
                i.speed_status,
                i.result_status,
                i.sampel_gratis,
                i.note,
                i.id_pesanan,
                i.no_va_co_sampel,
                i.status_payment_sampel,
                i.gmv_week1_after_join,
                i.gmv_week2_after_join,
                i.gmv_week3_after_join,
                i.gmv_week4_after_join,
                i.gmv_perbulan_after_join
            FROM influencers i
            LEFT JOIN brands b ON i.brand_id = b.id
            WHERE i.brand_id = :brand_id
            ORDER BY i.created_at DESC
        """
        result = await db.execute(text(query), {"brand_id": str(brand_row[0])})
        rows = result.fetchall()

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def generate_excel2_deal(
    db: AsyncSession,
    brand_id: Optional[str] = None,
) -> bytes:
    """Generate Excel 2 — Database Deal (sheet per brand)."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl tidak terinstall. Jalankan: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Ambil semua brand yang punya deal
    brand_query = """
        SELECT DISTINCT b.id, b.name
        FROM deal_records dr
        JOIN brands b ON dr.brand_id = b.id
        ORDER BY b.name
    """
    if brand_id:
        brand_query = """
            SELECT id, name FROM brands WHERE id = :brand_id
        """
    brand_result = await db.execute(text(brand_query), {"brand_id": brand_id} if brand_id else {})
    brands = brand_result.fetchall()

    if not brands:
        ws = wb.create_sheet("DEAL")
        _apply_header_style(ws, EXCEL2_HEADERS, "00B050")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    for brand_row in brands:
        sheet_name = f"{brand_row[1]} DEAL"[:31]
        ws = wb.create_sheet(sheet_name)
        _apply_header_style(ws, EXCEL2_HEADERS, "00B050")

        query = """
            SELECT
                dr.tanggal, dr.username, dr.link_acc, dr.follower_count,
                dr.contact_wa, b.name as brand, dr.pic,
                dr.avg_gmv_per_month, dr.gmv_per_buyer,
                dr.update_status, dr.respon_status, dr.speed_status, dr.result_status,
                dr.status_sempel, dr.link_video, dr.total_vt,
                dr.note_deal, dr.note_dari_rara,
                dr.id_pesanan, dr.no_va_co_sampel, dr.status_payment_sampel,
                dr.gmv_week1_after_join, dr.gmv_week2_after_join,
                dr.gmv_week3_after_join, dr.gmv_week4_after_join,
                dr.gmv_perbulan_after_join
            FROM deal_records dr
            JOIN brands b ON dr.brand_id = b.id
            WHERE dr.brand_id = :brand_id
            ORDER BY dr.tanggal DESC
        """
        result = await db.execute(text(query), {"brand_id": str(brand_row[0])})
        rows = result.fetchall()

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def generate_excel3_master_brand(db: AsyncSession) -> bytes:
    """Generate Excel 3 — Master Brand (SKU, WA, SOW per brand)."""
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl tidak terinstall. Jalankan: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOW DAN LINK SKU"
    _apply_header_style(ws, EXCEL3_HEADERS, "4472C4")

    query = """
        SELECT
            b.name as brand,
            bs.product_name,
            bs.affiliate_link,
            bs.price,
            b.wa_number,
            b.sow
        FROM brands b
        LEFT JOIN brand_skus bs ON b.id = bs.brand_id AND bs.is_active = TRUE
        WHERE b.is_active = TRUE
        ORDER BY b.name, bs.product_name
    """
    result = await db.execute(text(query))
    rows = result.fetchall()

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
