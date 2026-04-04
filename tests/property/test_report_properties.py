"""Property-based tests untuk ReportService dan NotificationService.

Validates: Requirements 10.1, 10.3, 14.2, 14.4
"""

from __future__ import annotations

import asyncio
import csv
import io
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
from unittest.mock import AsyncMock, MagicMock

import pytest
from hypothesis import given, settings
from hypothesis import strategies as st

from app.services.notification_service import Notification, NotificationService
from app.services.report_service import CampaignReportData, ReportService


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _run_async(coro):
    """Jalankan coroutine dalam event loop baru (kompatibel dengan Hypothesis)."""
    loop = asyncio.new_event_loop()
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()


def _make_db(
    inv_count: int = 10,
    acc_count: int = 5,
    total_views: int = 10000,
    total_gmv: float = 500000.0,
    avg_conversion: float = 0.05,
) -> AsyncMock:
    db = AsyncMock()
    db.flush = AsyncMock()

    async def _execute(query, params=None):
        q = str(query)
        mock_result = MagicMock()
        mock_mappings = MagicMock()

        if "COUNT(*) AS cnt FROM invitations" in q and "status = 'DELIVERED'" not in q:
            mock_mappings.first.return_value = {"cnt": inv_count}
        elif "COUNT(*) AS cnt FROM invitations" in q and "status = 'DELIVERED'" in q:
            mock_mappings.first.return_value = {"cnt": acc_count}
        elif "content_metrics" in q:
            mock_mappings.first.return_value = {
                "total_views": total_views,
                "total_gmv": total_gmv,
                "avg_conversion": avg_conversion,
            }
        else:
            mock_mappings.first.return_value = None

        mock_mappings.all.return_value = []
        mock_result.mappings.return_value = mock_mappings
        return mock_result

    db.execute = _execute
    return db


_EXPECTED_CSV_HEADERS = [
    "campaign_id", "total_influencers", "acceptance_rate",
    "total_views", "total_gmv", "cost_per_conversion", "generated_at",
]


# ---------------------------------------------------------------------------
# Property 32: Kelengkapan Laporan
# ---------------------------------------------------------------------------


class TestProperty32ReportCompleteness:
    """Validates: Requirements 10.1 — generate_campaign_report selalu mengandung semua field wajib
    dengan nilai valid (tidak None, tidak negatif)."""

    @given(
        inv_count=st.integers(min_value=0, max_value=1000),
        acc_count=st.integers(min_value=0, max_value=1000),
    )
    @settings(max_examples=50)
    def test_report_all_fields_not_none(self, inv_count: int, acc_count: int):
        """generate_campaign_report selalu mengandung semua field tidak None."""
        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(inv_count=inv_count, acc_count=actual_acc)
            service = ReportService()

            report = await service.generate_campaign_report("camp-1", db)

            assert report.campaign_id is not None
            assert report.total_influencers is not None
            assert report.acceptance_rate is not None
            assert report.total_views is not None
            assert report.total_gmv is not None
            assert report.cost_per_conversion is not None
            assert report.generated_at is not None

            assert isinstance(report, CampaignReportData)

        _run_async(_run())

    @given(
        inv_count=st.integers(min_value=0, max_value=1000),
        acc_count=st.integers(min_value=0, max_value=1000),
        total_views=st.integers(min_value=0, max_value=10_000_000),
        total_gmv=st.floats(min_value=0.0, max_value=1_000_000_000.0, allow_nan=False, allow_infinity=False),
        avg_conversion=st.floats(min_value=0.0, max_value=1.0, allow_nan=False, allow_infinity=False),
    )
    @settings(max_examples=50)
    def test_report_contains_all_required_metrics(
        self,
        inv_count: int,
        acc_count: int,
        total_views: int,
        total_gmv: float,
        avg_conversion: float,
    ):
        """Laporan selalu mengandung semua metrik wajib: total_influencers, acceptance_rate,
        total_views, total_gmv, cost_per_conversion (Requirement 10.1)."""
        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(
                inv_count=inv_count,
                acc_count=actual_acc,
                total_views=total_views,
                total_gmv=total_gmv,
                avg_conversion=avg_conversion,
            )
            service = ReportService()

            report = await service.generate_campaign_report("camp-req10", db)

            # Semua metrik wajib harus ada dan valid
            assert isinstance(report.total_influencers, int) and report.total_influencers >= 0
            assert isinstance(report.acceptance_rate, float) and 0.0 <= report.acceptance_rate <= 1.0
            assert isinstance(report.total_views, int) and report.total_views >= 0
            assert isinstance(report.total_gmv, float) and report.total_gmv >= 0.0
            assert isinstance(report.cost_per_conversion, float) and report.cost_per_conversion >= 0.0
            assert isinstance(report.generated_at, datetime)

        _run_async(_run())


# ---------------------------------------------------------------------------
# Property 33: Filter Laporan Konsisten
# ---------------------------------------------------------------------------


class TestProperty33FilterConsistent:
    """Validates: Requirements 10.3 — semua data dalam laporan yang difilter memenuhi kriteria filter.

    Karena generate_campaign_report menerima filter (start_date, end_date, categories, status),
    laporan yang dihasilkan harus konsisten: acceptance_rate dalam [0.0, 1.0], metrik numerik
    tidak negatif, dan campaign_id dalam laporan sama dengan campaign_id yang diminta.
    """

    @given(
        inv_count=st.integers(min_value=0, max_value=1000),
        acc_count=st.integers(min_value=0, max_value=1000),
    )
    @settings(max_examples=50)
    def test_acceptance_rate_in_0_to_1(self, inv_count: int, acc_count: int):
        """acceptance_rate selalu dalam [0.0, 1.0] untuk semua kombinasi inv/acc."""
        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(inv_count=inv_count, acc_count=actual_acc)
            service = ReportService()

            report = await service.generate_campaign_report("camp-1", db)

            assert 0.0 <= report.acceptance_rate <= 1.0, (
                f"acceptance_rate={report.acceptance_rate} harus dalam [0.0, 1.0] "
                f"untuk inv_count={inv_count}, acc_count={actual_acc}"
            )

        _run_async(_run())

    @given(
        campaign_id=st.text(min_size=1, max_size=50, alphabet=st.characters(whitelist_categories=("Lu", "Ll", "Nd"), whitelist_characters="-_")),
        inv_count=st.integers(min_value=0, max_value=500),
        acc_count=st.integers(min_value=0, max_value=500),
        total_views=st.integers(min_value=0, max_value=10_000_000),
        total_gmv=st.floats(min_value=0.0, max_value=1_000_000_000.0, allow_nan=False, allow_infinity=False),
    )
    @settings(max_examples=50)
    def test_filtered_report_campaign_id_matches_request(
        self,
        campaign_id: str,
        inv_count: int,
        acc_count: int,
        total_views: int,
        total_gmv: float,
    ):
        """campaign_id dalam laporan selalu sama dengan campaign_id yang diminta (filter konsisten)."""
        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(
                inv_count=inv_count,
                acc_count=actual_acc,
                total_views=total_views,
                total_gmv=total_gmv,
            )
            service = ReportService()

            start = datetime(2024, 1, 1, tzinfo=timezone.utc)
            end = datetime(2024, 12, 31, tzinfo=timezone.utc)
            report = await service.generate_campaign_report(
                campaign_id,
                db,
                start_date=start,
                end_date=end,
                categories=["fashion"],
                status="ACTIVE",
            )

            assert report.campaign_id == campaign_id, (
                f"campaign_id dalam laporan '{report.campaign_id}' "
                f"harus sama dengan yang diminta '{campaign_id}'"
            )

        _run_async(_run())

    @given(
        inv_count=st.integers(min_value=0, max_value=1000),
        acc_count=st.integers(min_value=0, max_value=1000),
        total_views=st.integers(min_value=0, max_value=10_000_000),
        total_gmv=st.floats(min_value=0.0, max_value=1_000_000_000.0, allow_nan=False, allow_infinity=False),
        avg_conversion=st.floats(min_value=0.0, max_value=1.0, allow_nan=False, allow_infinity=False),
    )
    @settings(max_examples=50)
    def test_filtered_report_numeric_metrics_non_negative(
        self,
        inv_count: int,
        acc_count: int,
        total_views: int,
        total_gmv: float,
        avg_conversion: float,
    ):
        """Semua metrik numerik dalam laporan yang difilter tidak boleh negatif."""
        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(
                inv_count=inv_count,
                acc_count=actual_acc,
                total_views=total_views,
                total_gmv=total_gmv,
                avg_conversion=avg_conversion,
            )
            service = ReportService()

            report = await service.generate_campaign_report(
                "camp-filter",
                db,
                start_date=datetime(2024, 1, 1, tzinfo=timezone.utc),
                end_date=datetime(2024, 12, 31, tzinfo=timezone.utc),
            )

            assert report.total_influencers >= 0, (
                f"total_influencers={report.total_influencers} tidak boleh negatif"
            )
            assert report.total_views >= 0, (
                f"total_views={report.total_views} tidak boleh negatif"
            )
            assert report.total_gmv >= 0.0, (
                f"total_gmv={report.total_gmv} tidak boleh negatif"
            )
            assert report.cost_per_conversion >= 0.0, (
                f"cost_per_conversion={report.cost_per_conversion} tidak boleh negatif"
            )

        _run_async(_run())


# ---------------------------------------------------------------------------
# Property 20: Notifikasi Threshold
# ---------------------------------------------------------------------------


class TestProperty20NotificationThreshold:
    """Validates: Requirements 14.4 — check_and_notify menghasilkan notifikasi jika metric > threshold."""

    @given(
        metric_value=st.floats(min_value=0.0, max_value=1_000_000.0, allow_nan=False, allow_infinity=False),
        threshold=st.floats(min_value=0.0, max_value=1_000_000.0, allow_nan=False, allow_infinity=False),
    )
    @settings(max_examples=50)
    def test_notification_when_metric_exceeds_threshold(
        self,
        metric_value: float,
        threshold: float,
    ):
        """For any metric > threshold, check_and_notify menghasilkan notifikasi."""
        async def _run():
            service = NotificationService()

            notifications = await service.check_and_notify(
                campaign_id="camp-1",
                metrics={"views": metric_value},
                alert_thresholds={"views": threshold},
            )

            if metric_value > threshold:
                assert len(notifications) == 1, (
                    f"Harus ada 1 notifikasi untuk metric={metric_value} > threshold={threshold}"
                )
                assert notifications[0].metric_name == "views"
                assert notifications[0].current_value == metric_value
            else:
                assert len(notifications) == 0, (
                    f"Tidak boleh ada notifikasi untuk metric={metric_value} <= threshold={threshold}"
                )

        _run_async(_run())

    @given(
        metric_value=st.floats(min_value=0.0, max_value=1_000_000.0, allow_nan=False, allow_infinity=False),
        threshold=st.floats(min_value=0.0, max_value=1_000_000.0, allow_nan=False, allow_infinity=False),
    )
    @settings(max_examples=50)
    def test_no_notification_when_metric_at_or_below_threshold(
        self,
        metric_value: float,
        threshold: float,
    ):
        """For any metric <= threshold, tidak ada notifikasi."""
        async def _run():
            service = NotificationService()

            # Pastikan metric <= threshold
            actual_metric = min(metric_value, threshold)

            notifications = await service.check_and_notify(
                campaign_id="camp-1",
                metrics={"views": actual_metric},
                alert_thresholds={"views": threshold},
            )

            assert len(notifications) == 0, (
                f"Tidak boleh ada notifikasi untuk metric={actual_metric} <= threshold={threshold}"
            )

        _run_async(_run())


# ---------------------------------------------------------------------------
# Property 21: Ekspor Data Lengkap (Requirements 6.5)
# ---------------------------------------------------------------------------


class TestProperty21ExportDataComplete:
    """Validates: Requirements 6.5 — file CSV/Excel yang dihasilkan mengandung semua baris
    dan kolom sesuai data di sistem."""

    # --- CSV ---

    @given(
        inv_count=st.integers(min_value=0, max_value=1000),
        acc_count=st.integers(min_value=0, max_value=1000),
    )
    @settings(max_examples=50)
    def test_export_csv_has_correct_headers(self, inv_count: int, acc_count: int):
        """export_csv selalu mengandung semua kolom wajib sesuai data di sistem."""
        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(inv_count=inv_count, acc_count=actual_acc)
            service = ReportService()

            csv_content = await service.export_csv("camp-1", db)

            reader = csv.reader(io.StringIO(csv_content))
            headers = next(reader)

            for expected_header in _EXPECTED_CSV_HEADERS:
                assert expected_header in headers, (
                    f"Kolom '{expected_header}' tidak ada dalam CSV. Kolom ada: {headers}"
                )

        _run_async(_run())

    @given(
        inv_count=st.integers(min_value=0, max_value=1000),
        acc_count=st.integers(min_value=0, max_value=1000),
    )
    @settings(max_examples=50)
    def test_export_csv_has_data_row(self, inv_count: int, acc_count: int):
        """export_csv harus mengandung setidaknya satu baris data (tidak hanya header)."""
        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(inv_count=inv_count, acc_count=actual_acc)
            service = ReportService()

            csv_content = await service.export_csv("camp-1", db)

            reader = csv.reader(io.StringIO(csv_content))
            rows = list(reader)

            # Minimal 2 baris: header + 1 data
            assert len(rows) >= 2, (
                f"CSV harus mengandung minimal 2 baris (header + data), tapi hanya {len(rows)}"
            )

        _run_async(_run())

    @given(
        campaign_id=st.text(
            min_size=1, max_size=50,
            alphabet=st.characters(whitelist_categories=("Lu", "Ll", "Nd"), whitelist_characters="-_"),
        ),
        inv_count=st.integers(min_value=0, max_value=500),
        acc_count=st.integers(min_value=0, max_value=500),
        total_views=st.integers(min_value=0, max_value=10_000_000),
        total_gmv=st.floats(min_value=0.0, max_value=1_000_000_000.0, allow_nan=False, allow_infinity=False),
        avg_conversion=st.floats(min_value=0.0, max_value=1.0, allow_nan=False, allow_infinity=False),
    )
    @settings(max_examples=50)
    def test_export_csv_data_matches_report(
        self,
        campaign_id: str,
        inv_count: int,
        acc_count: int,
        total_views: int,
        total_gmv: float,
        avg_conversion: float,
    ):
        """Nilai dalam baris data CSV harus konsisten dengan data laporan yang dihasilkan sistem
        (semua baris dan kolom sesuai data di sistem — Property 21)."""
        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(
                inv_count=inv_count,
                acc_count=actual_acc,
                total_views=total_views,
                total_gmv=total_gmv,
                avg_conversion=avg_conversion,
            )
            service = ReportService()

            # Ambil laporan langsung sebagai ground truth
            report = await service.generate_campaign_report(campaign_id, db)

            # Reset mock agar export_csv menggunakan data yang sama
            db2 = _make_db(
                inv_count=inv_count,
                acc_count=actual_acc,
                total_views=total_views,
                total_gmv=total_gmv,
                avg_conversion=avg_conversion,
            )
            csv_content = await service.export_csv(campaign_id, db2)

            reader = csv.DictReader(io.StringIO(csv_content))
            rows = list(reader)

            assert len(rows) >= 1, "CSV harus mengandung minimal satu baris data"

            row = rows[0]
            assert row["campaign_id"] == campaign_id, (
                f"campaign_id di CSV '{row['campaign_id']}' != '{campaign_id}'"
            )
            assert int(row["total_influencers"]) == inv_count, (
                f"total_influencers di CSV '{row['total_influencers']}' != '{inv_count}'"
            )
            assert int(row["total_views"]) == total_views, (
                f"total_views di CSV '{row['total_views']}' != '{total_views}'"
            )

        _run_async(_run())

    @given(
        inv_count=st.integers(min_value=0, max_value=1000),
        acc_count=st.integers(min_value=0, max_value=1000),
    )
    @settings(max_examples=50)
    def test_export_csv_column_count_matches_headers(self, inv_count: int, acc_count: int):
        """Setiap baris data CSV harus memiliki jumlah kolom yang sama dengan baris header."""
        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(inv_count=inv_count, acc_count=actual_acc)
            service = ReportService()

            csv_content = await service.export_csv("camp-1", db)

            reader = csv.reader(io.StringIO(csv_content))
            rows = list(reader)
            assert len(rows) >= 2

            header_count = len(rows[0])
            for i, row in enumerate(rows[1:], start=2):
                assert len(row) == header_count, (
                    f"Baris {i} memiliki {len(row)} kolom, tapi header memiliki {header_count} kolom"
                )

        _run_async(_run())

    # --- Excel ---

    @given(
        inv_count=st.integers(min_value=0, max_value=1000),
        acc_count=st.integers(min_value=0, max_value=1000),
    )
    @settings(max_examples=30, deadline=None)
    def test_export_excel_has_correct_headers(self, inv_count: int, acc_count: int):
        """export_excel selalu mengandung semua kolom wajib di baris pertama worksheet."""
        import openpyxl  # type: ignore

        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(inv_count=inv_count, acc_count=actual_acc)
            service = ReportService()

            excel_bytes = await service.export_excel("camp-1", db)

            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            for expected_header in _EXPECTED_CSV_HEADERS:
                assert expected_header in headers, (
                    f"Kolom '{expected_header}' tidak ada dalam Excel. Kolom ada: {headers}"
                )

        _run_async(_run())

    @given(
        inv_count=st.integers(min_value=0, max_value=1000),
        acc_count=st.integers(min_value=0, max_value=1000),
    )
    @settings(max_examples=30, deadline=None)
    def test_export_excel_has_data_row(self, inv_count: int, acc_count: int):
        """export_excel harus mengandung setidaknya satu baris data selain header."""
        import openpyxl  # type: ignore

        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(inv_count=inv_count, acc_count=actual_acc)
            service = ReportService()

            excel_bytes = await service.export_excel("camp-1", db)

            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
            ws = wb.active
            row_count = ws.max_row

            assert row_count >= 2, (
                f"Excel harus mengandung minimal 2 baris (header + data), tapi hanya {row_count}"
            )

        _run_async(_run())

    @given(
        campaign_id=st.text(
            min_size=1, max_size=50,
            alphabet=st.characters(whitelist_categories=("Lu", "Ll", "Nd"), whitelist_characters="-_"),
        ),
        inv_count=st.integers(min_value=0, max_value=500),
        acc_count=st.integers(min_value=0, max_value=500),
        total_views=st.integers(min_value=0, max_value=10_000_000),
        total_gmv=st.floats(min_value=0.0, max_value=1_000_000_000.0, allow_nan=False, allow_infinity=False),
        avg_conversion=st.floats(min_value=0.0, max_value=1.0, allow_nan=False, allow_infinity=False),
    )
    @settings(max_examples=30, deadline=None)
    def test_export_excel_data_matches_report(
        self,
        campaign_id: str,
        inv_count: int,
        acc_count: int,
        total_views: int,
        total_gmv: float,
        avg_conversion: float,
    ):
        """Nilai dalam baris data Excel harus konsisten dengan data laporan yang dihasilkan sistem
        (semua baris dan kolom sesuai data di sistem — Property 21)."""
        import openpyxl  # type: ignore

        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(
                inv_count=inv_count,
                acc_count=actual_acc,
                total_views=total_views,
                total_gmv=total_gmv,
                avg_conversion=avg_conversion,
            )
            service = ReportService()

            excel_bytes = await service.export_excel(campaign_id, db)

            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
            ws = wb.active

            # Baris 1 = header, baris 2 = data
            headers = [cell.value for cell in ws[1]]
            data_row = [cell.value for cell in ws[2]]
            row_dict = dict(zip(headers, data_row))

            assert row_dict["campaign_id"] == campaign_id, (
                f"campaign_id di Excel '{row_dict['campaign_id']}' != '{campaign_id}'"
            )
            assert int(row_dict["total_influencers"]) == inv_count, (
                f"total_influencers di Excel '{row_dict['total_influencers']}' != '{inv_count}'"
            )
            assert int(row_dict["total_views"]) == total_views, (
                f"total_views di Excel '{row_dict['total_views']}' != '{total_views}'"
            )

        _run_async(_run())

    @given(
        inv_count=st.integers(min_value=0, max_value=1000),
        acc_count=st.integers(min_value=0, max_value=1000),
    )
    @settings(max_examples=30, deadline=None)
    def test_export_excel_column_count_matches_headers(self, inv_count: int, acc_count: int):
        """Setiap baris data Excel harus memiliki jumlah kolom yang sama dengan baris header."""
        import openpyxl  # type: ignore

        async def _run():
            actual_acc = min(acc_count, inv_count)
            db = _make_db(inv_count=inv_count, acc_count=actual_acc)
            service = ReportService()

            excel_bytes = await service.export_excel("camp-1", db)

            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
            ws = wb.active

            header_count = ws.max_column
            for row_idx in range(2, ws.max_row + 1):
                row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, header_count + 1)]
                non_none = [v for v in row_values if v is not None]
                assert len(non_none) == header_count, (
                    f"Baris {row_idx} di Excel memiliki {len(non_none)} nilai non-null, "
                    f"tapi header memiliki {header_count} kolom"
                )

        _run_async(_run())
