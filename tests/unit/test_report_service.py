"""Unit tests for ReportService."""

from __future__ import annotations

from datetime import datetime, timezone
from typing import Any, Dict, Optional
from unittest.mock import AsyncMock, MagicMock

import pytest

from app.services.report_service import CampaignReportData, ReportService


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_db(
    inv_count: int = 5,
    acc_count: int = 3,
    total_views: int = 10_000,
    total_gmv: float = 500_000.0,
    avg_conversion: float = 0.05,
) -> AsyncMock:
    """Build a mock AsyncSession that returns preset aggregate values."""

    def _make_result(mapping: Dict[str, Any]) -> MagicMock:
        row = MagicMock()
        row.__getitem__ = lambda self, key: mapping[key]
        row.get = lambda key, default=None: mapping.get(key, default)

        mappings_mock = MagicMock()
        mappings_mock.first = MagicMock(return_value=row)

        result = MagicMock()
        result.mappings = MagicMock(return_value=mappings_mock)
        return result

    inv_result = _make_result({"cnt": inv_count})
    acc_result = _make_result({"cnt": acc_count})
    metrics_result = _make_result({
        "total_views": total_views,
        "total_gmv": total_gmv,
        "avg_conversion": avg_conversion,
    })

    db = AsyncMock()
    db.execute = AsyncMock(side_effect=[inv_result, acc_result, metrics_result])
    return db


def _make_service() -> ReportService:
    return ReportService()


# ---------------------------------------------------------------------------
# Tests: generate_campaign_report
# ---------------------------------------------------------------------------


class TestGenerateCampaignReport:
    @pytest.mark.asyncio
    async def test_returns_campaign_report_data_instance(self):
        db = _make_db()
        service = _make_service()
        report = await service.generate_campaign_report("camp-1", db)
        assert isinstance(report, CampaignReportData)

    @pytest.mark.asyncio
    async def test_campaign_id_matches(self):
        db = _make_db()
        service = _make_service()
        report = await service.generate_campaign_report("camp-42", db)
        assert report.campaign_id == "camp-42"

    @pytest.mark.asyncio
    async def test_total_influencers_correct(self):
        db = _make_db(inv_count=10)
        service = _make_service()
        report = await service.generate_campaign_report("camp-1", db)
        assert report.total_influencers == 10

    @pytest.mark.asyncio
    async def test_acceptance_rate_correct(self):
        db = _make_db(inv_count=10, acc_count=4)
        service = _make_service()
        report = await service.generate_campaign_report("camp-1", db)
        assert report.acceptance_rate == pytest.approx(0.4)

    @pytest.mark.asyncio
    async def test_acceptance_rate_zero_when_no_invitations(self):
        db = _make_db(inv_count=0, acc_count=0)
        service = _make_service()
        report = await service.generate_campaign_report("camp-1", db)
        assert report.acceptance_rate == 0.0

    @pytest.mark.asyncio
    async def test_total_views_correct(self):
        db = _make_db(total_views=99_999)
        service = _make_service()
        report = await service.generate_campaign_report("camp-1", db)
        assert report.total_views == 99_999

    @pytest.mark.asyncio
    async def test_total_gmv_correct(self):
        db = _make_db(total_gmv=1_234_567.89)
        service = _make_service()
        report = await service.generate_campaign_report("camp-1", db)
        assert report.total_gmv == pytest.approx(1_234_567.89)

    @pytest.mark.asyncio
    async def test_cost_per_conversion_correct(self):
        # cost_per_conversion = total_gmv / avg_conversion
        db = _make_db(total_gmv=500_000.0, avg_conversion=0.1)
        service = _make_service()
        report = await service.generate_campaign_report("camp-1", db)
        assert report.cost_per_conversion == pytest.approx(5_000_000.0)

    @pytest.mark.asyncio
    async def test_cost_per_conversion_zero_when_no_conversion(self):
        db = _make_db(total_gmv=500_000.0, avg_conversion=0.0)
        service = _make_service()
        report = await service.generate_campaign_report("camp-1", db)
        assert report.cost_per_conversion == 0.0

    @pytest.mark.asyncio
    async def test_generated_at_is_datetime(self):
        db = _make_db()
        service = _make_service()
        report = await service.generate_campaign_report("camp-1", db)
        assert isinstance(report.generated_at, datetime)

    @pytest.mark.asyncio
    async def test_all_metrics_present(self):
        db = _make_db()
        service = _make_service()
        report = await service.generate_campaign_report("camp-1", db)
        # All required fields must be present and not None
        assert report.total_influencers is not None
        assert report.acceptance_rate is not None
        assert report.total_views is not None
        assert report.total_gmv is not None
        assert report.cost_per_conversion is not None


# ---------------------------------------------------------------------------
# Tests: export_csv
# ---------------------------------------------------------------------------


class TestExportCsv:
    @pytest.mark.asyncio
    async def test_export_csv_returns_string(self):
        db = _make_db()
        service = _make_service()
        result = await service.export_csv("camp-1", db)
        assert isinstance(result, str)

    @pytest.mark.asyncio
    async def test_export_csv_contains_header(self):
        db = _make_db()
        service = _make_service()
        result = await service.export_csv("camp-1", db)
        assert "campaign_id" in result
        assert "total_influencers" in result
        assert "acceptance_rate" in result
        assert "total_views" in result
        assert "total_gmv" in result
        assert "cost_per_conversion" in result

    @pytest.mark.asyncio
    async def test_export_csv_contains_campaign_id(self):
        db = _make_db()
        service = _make_service()
        result = await service.export_csv("camp-xyz", db)
        assert "camp-xyz" in result

    @pytest.mark.asyncio
    async def test_export_csv_has_two_lines(self):
        db = _make_db()
        service = _make_service()
        result = await service.export_csv("camp-1", db)
        lines = [l for l in result.strip().splitlines() if l]
        assert len(lines) == 2  # header + data row


# ---------------------------------------------------------------------------
# Tests: export_excel
# ---------------------------------------------------------------------------


class TestExportExcel:
    @pytest.mark.asyncio
    async def test_export_excel_returns_bytes(self):
        db = _make_db()
        service = _make_service()
        result = await service.export_excel("camp-1", db)
        assert isinstance(result, bytes)

    @pytest.mark.asyncio
    async def test_export_excel_is_valid_xlsx(self):
        import openpyxl
        import io

        db = _make_db()
        service = _make_service()
        result = await service.export_excel("camp-1", db)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        # Header row — accept both English and Indonesian headers
        headers = [cell.value for cell in ws[1]]
        assert any(h in headers for h in ["campaign_id", "ID Kampanye"]), f"No campaign ID header found in {headers}"
        assert any(h in headers for h in ["total_influencers", "Total Influencer"]), f"No influencer count header found in {headers}"

    @pytest.mark.asyncio
    async def test_export_excel_data_row_present(self):
        import openpyxl
        import io

        db = _make_db(inv_count=7)
        service = _make_service()
        result = await service.export_excel("camp-1", db)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        assert ws.max_row == 2  # header + 1 data row


# ---------------------------------------------------------------------------
# Tests: export_pdf
# ---------------------------------------------------------------------------


class TestExportPdf:
    @pytest.mark.asyncio
    async def test_export_pdf_returns_bytes(self):
        db = _make_db()
        service = _make_service()
        result = await service.export_pdf("camp-1", db)
        assert isinstance(result, bytes)

    @pytest.mark.asyncio
    async def test_export_pdf_starts_with_pdf_magic(self):
        db = _make_db()
        service = _make_service()
        result = await service.export_pdf("camp-1", db)
        assert result[:4] == b"%PDF"
